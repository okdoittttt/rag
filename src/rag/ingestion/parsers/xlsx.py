"""Excel(.xlsx) 파서

openpyxl을 사용하여 .xlsx 파일의 각 시트를 행 단위 텍스트로 변환합니다.
"""

from __future__ import annotations

from pathlib import Path

from rag.ingestion.parsers.base import DocumentParser
from rag.logger import get_logger


logger = get_logger(__name__)


class XlsxParser(DocumentParser):
    """Excel(.xlsx) 파일 파서"""

    extensions = [".xlsx"]

    def parse(self, path: Path) -> str:
        """.xlsx 파일에서 텍스트를 추출한다.

        시트별로 ``# {시트명}`` 머리말을 붙이고, 각 행의 비어있지 않은 셀 값을
        ``" | "``로 연결한다. 수식 셀은 계산된 값(``data_only=True``)을 사용하며,
        메모리 사용을 줄이기 위해 ``read_only`` 모드로 연다.

        Args:
            path: 파싱할 .xlsx 파일 경로.

        Returns:
            추출된 텍스트. 추출 실패 시 빈 문자열.

        Raises:
            ValueError: ``openpyxl`` 이 설치되어 있지 않을 때 발생.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            logger.warning("xlsx_dependency_missing", error=str(e))
            raise ValueError(
                "openpyxl이 설치되어 있지 않습니다. 'uv sync'로 의존성을 설치하세요."
            ) from e

        try:
            workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            logger.warning("xlsx_parse_failed", path=str(path), error=str(e))
            return ""

        try:
            parts: list[str] = []
            for sheet in workbook.worksheets:
                rows: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [
                        str(value).strip()
                        for value in row
                        if value is not None and str(value).strip()
                    ]
                    if cells:
                        rows.append(" | ".join(cells))
                if rows:
                    parts.append(f"# {sheet.title}\n" + "\n".join(rows))
            return "\n\n".join(parts)
        finally:
            workbook.close()
